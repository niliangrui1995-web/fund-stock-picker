from __future__ import annotations

import argparse
import ast
import csv
import json
import re
import sys
import threading
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

from quarter_config import cutoff_date_for_quarter, load_quarter_config, month_for_quarter, report_label
from spreadsheet_safety import append_safe_xlsx_row, safe_csv_row, safe_xlsx_value


FUND_LIST_URL = "https://fund.eastmoney.com/js/fundcode_search.js"
HOLDING_URL = (
    "https://fundf10.eastmoney.com/FundArchivesDatas.aspx"
    "?type={api_type}&code={fund_code}&topline={topline}&year={year}&month={month}"
)
REFERER_URL = "https://fundf10.eastmoney.com/ccmx_{fund_code}.html"

TYPE_CONFIG = {
    "stock": {
        "api_type": "jjcc",
        "label": "股票",
        "sheet": "股票持仓",
        "code_header": "股票代码",
        "name_header": "股票名称",
    },
    "bond": {
        "api_type": "zqcc",
        "label": "债券",
        "sheet": "债券持仓",
        "code_header": "债券代码",
        "name_header": "债券名称",
    },
}

HOLDING_HEADERS = [
    "基金代码",
    "基金名称",
    "基金类型",
    "报告期",
    "截止日期",
    "持仓类别",
    "序号",
    "证券代码",
    "证券名称",
    "占净值比例",
    "占净值比例数值",
    "持仓市值(万元)",
    "持股数(万股)",
    "最新价",
    "涨跌幅",
    "涨跌幅数值",
    "来源标题",
    "来源URL",
]

FUND_HEADERS = ["基金代码", "基金名称", "基金类型", "拼音缩写", "拼音全称", "是否QDII"]
NUMERIC_HEADERS = {"占净值比例数值", "持仓市值(万元)", "持股数(万股)", "最新价", "涨跌幅数值"}
MAX_EXCEL_ROWS = 1_048_576
DATA_ROWS_PER_SHEET = MAX_EXCEL_ROWS - 1

THREAD_LOCAL = threading.local()


def parse_args() -> argparse.Namespace:
    quarter_config = load_quarter_config()
    parser = argparse.ArgumentParser(
        description=f"Fetch Eastmoney/Tiantian fund holdings and build the configured {quarter_config.report} workbook."
    )
    parser.add_argument("--year", type=int, default=quarter_config.year)
    parser.add_argument("--quarter", type=int, default=quarter_config.quarter, choices=[1, 2, 3, 4])
    parser.add_argument(
        "--types",
        default="stock,bond",
        help="Comma-separated holding types: stock,bond",
    )
    parser.add_argument("--workers", type=int, default=10)
    parser.add_argument("--topline", type=int, default=1000)
    parser.add_argument("--limit", type=int, default=0, help="Limit funds for a smoke test.")
    parser.add_argument("--refresh", action="store_true", help="Ignore cached endpoint responses.")
    parser.add_argument("--progress-every", type=int, default=500)
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument("--cache-dir", default="data/eastmoney_cache")
    return parser.parse_args()


def get_session() -> requests.Session:
    session = getattr(THREAD_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/125.0 Safari/537.36"
                ),
                "Accept": "*/*",
            }
        )
        THREAD_LOCAL.session = session
    return session


def fetch_text(url: str, cache_path: Path, refresh: bool, referer: str | None = None) -> tuple[str, str]:
    if cache_path.exists() and not refresh:
        return cache_path.read_text(encoding="utf-8", errors="replace"), "cache"

    headers = {}
    if referer:
        headers["Referer"] = referer

    last_error: Exception | None = None
    for attempt in range(4):
        try:
            response = get_session().get(url, headers=headers, timeout=25)
            response.raise_for_status()
            response.encoding = "utf-8"
            text = response.text
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            tmp_path = cache_path.with_suffix(cache_path.suffix + ".tmp")
            tmp_path.write_text(text, encoding="utf-8")
            tmp_path.replace(cache_path)
            return text, "network"
        except Exception as exc:  # requests raises several concrete subclasses here
            last_error = exc
            time.sleep(0.8 * (attempt + 1))

    raise RuntimeError(str(last_error))


def parse_fund_list(text: str) -> list[dict[str, str]]:
    match = re.search(r"var\s+r\s*=\s*(\[.*\])\s*;?\s*$", text, flags=re.S)
    if not match:
        raise ValueError("Could not parse Eastmoney fund list response.")
    rows = ast.literal_eval(match.group(1))
    funds = []
    for code, pinyin_short, name, fund_type, pinyin_full in rows:
        funds.append(
            {
                "code": str(code).zfill(6),
                "pinyin_short": pinyin_short,
                "name": name,
                "type": fund_type,
                "pinyin_full": pinyin_full,
                "is_qdii": "Y" if fund_type.startswith("QDII") else "",
            }
        )
    return funds


def clean_content(raw: str) -> str:
    return (
        raw.replace('\\"', '"')
        .replace("\\/", "/")
        .replace("\\r", "")
        .replace("\\n", "")
        .replace("\\t", "")
    )


def extract_content(text: str) -> str:
    if text.strip() == "var apidata=":
        return ""
    match = re.search(r'content:"(?P<content>.*?)",(?:arryear|curyear|records|pages)', text, flags=re.S)
    if not match:
        match = re.search(r'content:"(?P<content>.*)"\s*}', text, flags=re.S)
    if not match:
        return ""
    return clean_content(match.group("content"))


def canon_header(value: str) -> str:
    return re.sub(r"[\s\xa0]+", "", value or "")


def pct_number(value: str) -> float | None:
    value = (value or "").strip()
    if not value or value == "--":
        return None
    value = value.replace("%", "").replace(",", "")
    try:
        return float(value) / 100
    except ValueError:
        return None


def plain_number(value: str) -> float | None:
    value = (value or "").strip()
    if not value or value == "--":
        return None
    value = value.replace(",", "")
    try:
        return float(value)
    except ValueError:
        return None


def first_value(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(canon_header(key), "")
        if value:
            return value
    return ""


def parse_holding_rows(
    fund: dict[str, str],
    holding_type: str,
    text: str,
    source_url: str,
    year: int,
    quarter: int,
) -> dict[str, Any]:
    content = extract_content(text)
    if not content:
        return {"status": "no_data", "rows": [], "cutoff_date": "", "title": "", "error": ""}

    soup = BeautifulSoup(content, "html.parser")
    title = " ".join(node.get_text(" ", strip=True) for node in soup.select("h4.t"))
    title = re.sub(r"[\s\xa0]+", " ", title).strip()
    cutoff_match = re.search(r"\d{4}-\d{2}-\d{2}", title or soup.get_text(" ", strip=True))
    cutoff_date = cutoff_match.group(0) if cutoff_match else ""

    table = soup.find("table")
    if table is None:
        page_text = soup.get_text(" ", strip=True)
        if "暂无" in page_text or not page_text:
            return {"status": "no_data", "rows": [], "cutoff_date": cutoff_date, "title": title, "error": ""}
        return {
            "status": "parse_error",
            "rows": [],
            "cutoff_date": cutoff_date,
            "title": title,
            "error": "No table found in content.",
        }

    headers = [canon_header(th.get_text(" ", strip=True)) for th in table.find_all("th")]
    data_rows = []
    for tr in table.find_all("tr"):
        cells = [td.get_text(" ", strip=True) for td in tr.find_all("td")]
        if not cells:
            continue
        if len(cells) != len(headers):
            continue
        row = {headers[index]: cells[index] for index in range(len(headers))}
        nav_ratio = first_value(row, "占净值比例", "占净值 比例")
        change_pct = first_value(row, "涨跌幅")
        data_rows.append(
            [
                fund["code"],
                fund["name"],
                fund["type"],
                report_label(year, quarter),
                cutoff_date,
                TYPE_CONFIG[holding_type]["label"],
                first_value(row, "序号"),
                first_value(row, TYPE_CONFIG[holding_type]["code_header"]),
                first_value(row, TYPE_CONFIG[holding_type]["name_header"]),
                nav_ratio,
                pct_number(nav_ratio),
                plain_number(first_value(row, "持仓市值（万元）", "持仓市值(万元)")),
                plain_number(first_value(row, "持股数（万股）", "持股数(万股)")),
                plain_number(first_value(row, "最新价")),
                change_pct,
                pct_number(change_pct),
                title,
                source_url,
            ]
        )

    expected_cutoff_date = cutoff_date_for_quarter(year, quarter)
    if data_rows and cutoff_date != expected_cutoff_date:
        return {
            "status": "out_of_period",
            "rows": [],
            "cutoff_date": cutoff_date,
            "title": title,
            "error": f"非目标报告期：期望截止日 {expected_cutoff_date}，实际 {cutoff_date or '空'}。",
        }

    status = "ok" if data_rows else "no_data"
    return {"status": status, "rows": data_rows, "cutoff_date": cutoff_date, "title": title, "error": ""}


def fetch_one_fund(
    fund: dict[str, str],
    selected_types: list[str],
    cache_root: Path,
    refresh: bool,
    year: int,
    quarter: int,
    topline: int,
) -> dict[str, Any]:
    month = month_for_quarter(quarter)
    result: dict[str, Any] = {
        "fund": fund,
        "holdings": {holding_type: [] for holding_type in selected_types},
        "status": {},
    }
    for holding_type in selected_types:
        config = TYPE_CONFIG[holding_type]
        source_url = HOLDING_URL.format(
            api_type=config["api_type"],
            fund_code=fund["code"],
            topline=topline,
            year=year,
            month=month,
        )
        cache_path = cache_root / report_label(year, quarter) / config["api_type"] / f"{fund['code']}.txt"
        try:
            text, fetch_source = fetch_text(
                source_url,
                cache_path=cache_path,
                refresh=refresh,
                referer=REFERER_URL.format(fund_code=fund["code"]),
            )
            parsed = parse_holding_rows(fund, holding_type, text, source_url, year, quarter)
            result["holdings"][holding_type] = parsed["rows"]
            result["status"][holding_type] = {
                "status": parsed["status"],
                "rows": len(parsed["rows"]),
                "cutoff_date": parsed["cutoff_date"],
                "fetch_source": fetch_source,
                "error": parsed["error"],
            }
        except Exception as exc:
            result["status"][holding_type] = {
                "status": "error",
                "rows": 0,
                "cutoff_date": "",
                "fetch_source": "",
                "error": str(exc),
            }
    return result


def write_fund_list(funds: list[dict[str, str]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, quoting=csv.QUOTE_ALL)
        writer.writerow(safe_csv_row(FUND_HEADERS))
        for fund in funds:
            writer.writerow(
                safe_csv_row(
                    [
                        fund["code"],
                        fund["name"],
                        fund["type"],
                        fund["pinyin_short"],
                        fund["pinyin_full"],
                        fund["is_qdii"],
                    ]
                )
            )


def build_status_header(selected_types: list[str]) -> list[str]:
    header = ["基金代码", "基金名称", "基金类型"]
    for holding_type in selected_types:
        label = TYPE_CONFIG[holding_type]["label"]
        header.extend([f"{label}状态", f"{label}行数", f"{label}截止日期", f"{label}来源", f"{label}备注"])
    return header


def status_row(record: dict[str, Any], selected_types: list[str]) -> list[Any]:
    fund = record["fund"]
    row: list[Any] = [fund["code"], fund["name"], fund["type"]]
    for holding_type in selected_types:
        status = record["status"].get(holding_type, {})
        row.extend(
            [
                status.get("status", ""),
                status.get("rows", 0),
                status.get("cutoff_date", ""),
                status.get("fetch_source", ""),
                status.get("error", ""),
            ]
        )
    return row


def read_csv_row_count(path: Path) -> int:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return max(sum(1 for _ in handle) - 1, 0)


def make_header_cells(ws: Any, values: list[str]) -> list[WriteOnlyCell]:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(name="Arial", bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")
    cells = []
    for value in values:
        safe_value, is_formula = safe_xlsx_value(value)
        cell = WriteOnlyCell(ws, value=safe_value)
        if is_formula:
            cell.data_type = "s"
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cells.append(cell)
    return cells


def coerce_cell(header: str, value: str) -> Any:
    if value == "":
        return None
    if header in NUMERIC_HEADERS or header.endswith("行数"):
        try:
            return float(value)
        except ValueError:
            return value
    return value


def safe_sheet_name(base_name: str, index: int) -> str:
    suffix = f"_{index:02d}"
    return f"{base_name[:31 - len(suffix)]}{suffix}"


def append_csv_as_sheets(wb: Workbook, csv_path: Path, base_sheet_name: str) -> list[dict[str, Any]]:
    created = []
    with csv_path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        sheet_index = 1
        row_in_sheet = 0
        ws = wb.create_sheet(safe_sheet_name(base_sheet_name, sheet_index))
        ws.freeze_panes = "A2"
        ws.append(make_header_cells(ws, header))
        row_in_sheet = 1
        sheet_rows = 0

        for row in reader:
            if row_in_sheet >= MAX_EXCEL_ROWS:
                created.append({"sheet": ws.title, "data_rows": sheet_rows})
                sheet_index += 1
                ws = wb.create_sheet(safe_sheet_name(base_sheet_name, sheet_index))
                ws.freeze_panes = "A2"
                ws.append(make_header_cells(ws, header))
                row_in_sheet = 1
                sheet_rows = 0

            append_safe_xlsx_row(
                ws,
                [coerce_cell(header[i], row[i]) if i < len(row) else None for i in range(len(header))],
            )
            row_in_sheet += 1
            sheet_rows += 1

        created.append({"sheet": ws.title, "data_rows": sheet_rows})
    return created


def add_readme_sheet(
    wb: Workbook,
    *,
    year: int,
    quarter: int,
    selected_types: list[str],
    fund_count: int,
    summary: dict[str, Any],
) -> None:
    ws = wb.create_sheet("说明", 0)
    ws.freeze_panes = "A2"
    ws.append(make_header_cells(ws, ["项目", "内容"]))
    rows = [
        ["报告期", report_label(year, quarter)],
        ["季度截止日", cutoff_date_for_quarter(year, quarter)],
        ["基金清单来源", FUND_LIST_URL],
        ["持仓来源", "天天基金 / 东方财富基金F10 FundArchivesDatas.aspx"],
        ["持仓类别", ",".join(TYPE_CONFIG[item]["label"] for item in selected_types)],
        ["基金数量", fund_count],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["说明", "季度公开页面通常披露股票/债券投资明细；未返回表格的基金在拉取状态中标为 no_data。"],
        ["完整性摘要", json.dumps(summary, ensure_ascii=False, sort_keys=True)],
    ]
    for row in rows:
        append_safe_xlsx_row(ws, row)


def build_workbook(
    output_path: Path,
    fund_csv: Path,
    holding_csvs: dict[str, Path],
    status_csv: Path,
    year: int,
    quarter: int,
    selected_types: list[str],
    summary: dict[str, Any],
) -> list[dict[str, Any]]:
    wb = Workbook(write_only=True)
    add_readme_sheet(
        wb,
        year=year,
        quarter=quarter,
        selected_types=selected_types,
        fund_count=summary["fund_count"],
        summary=summary,
    )

    sheet_summaries = []
    sheet_summaries.extend(append_csv_as_sheets(wb, fund_csv, "基金清单"))
    for holding_type in selected_types:
        sheet_summaries.extend(append_csv_as_sheets(wb, holding_csvs[holding_type], TYPE_CONFIG[holding_type]["sheet"]))
    sheet_summaries.extend(append_csv_as_sheets(wb, status_csv, "拉取状态"))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return sheet_summaries


def validate_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return {
            "opened": True,
            "sheetnames": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
        }
    finally:
        wb.close()


def main() -> int:
    args = parse_args()
    selected_types = [item.strip() for item in args.types.split(",") if item.strip()]
    invalid_types = [item for item in selected_types if item not in TYPE_CONFIG]
    if invalid_types:
        raise ValueError(f"Unsupported holding types: {invalid_types}")

    root = Path.cwd()
    output_dir = root / args.output_dir
    cache_root = root / args.cache_dir
    label = report_label(args.year, args.quarter).lower()

    fund_list_cache = cache_root / label / "fundcode_search.js"
    fund_text, _ = fetch_text(FUND_LIST_URL, fund_list_cache, refresh=args.refresh)
    funds = parse_fund_list(fund_text)
    funds.sort(key=lambda item: item["code"])
    if args.limit:
        funds = funds[: args.limit]

    output_dir.mkdir(parents=True, exist_ok=True)
    fund_csv = output_dir / f"fund_list_{label}.csv"
    status_csv = output_dir / f"fetch_status_{label}.csv"
    holding_csvs = {
        holding_type: output_dir / f"holdings_{holding_type}_{label}.csv" for holding_type in selected_types
    }
    workbook_path = output_dir / f"fund_holdings_{label}.xlsx"
    summary_path = output_dir / f"run_summary_{label}.json"

    write_fund_list(funds, fund_csv)

    started = time.time()
    status_counts: dict[str, Counter[str]] = {holding_type: Counter() for holding_type in selected_types}
    processed = 0

    holding_handles = {
        holding_type: holding_csvs[holding_type].open("w", newline="", encoding="utf-8-sig")
        for holding_type in selected_types
    }
    try:
        holding_writers = {
            holding_type: csv.writer(handle, quoting=csv.QUOTE_ALL) for holding_type, handle in holding_handles.items()
        }
        for writer in holding_writers.values():
            writer.writerow(safe_csv_row(HOLDING_HEADERS))

        with status_csv.open("w", newline="", encoding="utf-8-sig") as status_handle:
            status_writer = csv.writer(status_handle, quoting=csv.QUOTE_ALL)
            status_writer.writerow(safe_csv_row(build_status_header(selected_types)))

            with ThreadPoolExecutor(max_workers=max(args.workers, 1)) as executor:
                futures = [
                    executor.submit(
                        fetch_one_fund,
                        fund,
                        selected_types,
                        cache_root,
                        args.refresh,
                        args.year,
                        args.quarter,
                        args.topline,
                    )
                    for fund in funds
                ]
                for future in as_completed(futures):
                    record = future.result()
                    processed += 1
                    for holding_type in selected_types:
                        rows = record["holdings"].get(holding_type, [])
                        holding_writers[holding_type].writerows(safe_csv_row(row) for row in rows)
                        status = record["status"].get(holding_type, {}).get("status", "missing")
                        status_counts[holding_type][status] += 1
                    status_writer.writerow(safe_csv_row(status_row(record, selected_types)))

                    if args.progress_every and processed % args.progress_every == 0:
                        elapsed = time.time() - started
                        print(
                            json.dumps(
                                {
                                    "processed": processed,
                                    "fund_count": len(funds),
                                    "elapsed_seconds": round(elapsed, 1),
                                    "status_counts": {k: dict(v) for k, v in status_counts.items()},
                                },
                                ensure_ascii=False,
                            ),
                            flush=True,
                        )
    finally:
        for handle in holding_handles.values():
            handle.close()

    holding_rows = {holding_type: read_csv_row_count(path) for holding_type, path in holding_csvs.items()}
    status_rows = read_csv_row_count(status_csv)
    summary: dict[str, Any] = {
        "report": report_label(args.year, args.quarter),
        "fund_count": len(funds),
        "status_rows": status_rows,
        "selected_types": selected_types,
        "holding_rows": holding_rows,
        "status_counts": {holding_type: dict(status_counts[holding_type]) for holding_type in selected_types},
        "fund_csv": str(fund_csv),
        "holding_csvs": {holding_type: str(path) for holding_type, path in holding_csvs.items()},
        "status_csv": str(status_csv),
        "workbook": str(workbook_path),
    }

    sheet_summaries = build_workbook(
        workbook_path,
        fund_csv,
        holding_csvs,
        status_csv,
        args.year,
        args.quarter,
        selected_types,
        summary,
    )
    summary["workbook_sheets"] = sheet_summaries
    summary["workbook_validation"] = validate_workbook(workbook_path)
    summary["elapsed_seconds"] = round(time.time() - started, 1)
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("Interrupted.", file=sys.stderr)
        raise SystemExit(130)
