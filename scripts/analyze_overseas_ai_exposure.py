from __future__ import annotations

import csv
import ast
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from quarter_config import load_quarter_config


QUARTER = load_quarter_config()
INPUT = QUARTER.source_stock_csv
OUTPUT_XLSX = QUARTER.overseas_ai_workbook
OUTPUT_RANKING_CSV = QUARTER.overseas_ai_ranking_csv
OUTPUT_DETAIL_CSV = QUARTER.overseas_ai_detail_csv
OUTPUT_PURCHASE_LIMIT_CSV = Path("outputs/fund_purchase_limit_snapshot.csv")
OUTPUT_SUMMARY_JSON = QUARTER.overseas_ai_summary_json
PURCHASE_LIMIT_URL = "https://fund.eastmoney.com/Data/Fund_JJJZ_Data.aspx"

STRICT_OVERSEAS_AI = {
    "NVDA": ("英伟达", "核心算力-GPU"),
    "AVGO": ("博通", "AI网络/ASIC"),
    "AMD": ("超威半导体", "核心算力-GPU"),
    "MU": ("美光科技", "AI存储/HBM"),
    "000660": ("SK海力士", "AI存储/HBM"),
    "TSM": ("台积电ADR", "先进制程代工"),
    "2330": ("台积电", "先进制程代工"),
    "ASML": ("阿斯麦", "半导体设备"),
    "MRVL": ("迈威尔科技", "AI网络/ASIC"),
    "AMAT": ("应用材料", "半导体设备"),
    "LRCX": ("拉姆研究", "半导体设备"),
    "KLAC": ("科磊", "半导体设备"),
    "ARM": ("Arm Holdings", "AI芯片IP"),
    "VRT": ("Vertiv", "AI数据中心电力/散热"),
    "MSFT": ("微软", "AI云/软件"),
    "GOOGL": ("谷歌A", "AI云/大模型"),
    "GOOG": ("谷歌C", "AI云/大模型"),
    "AMZN": ("亚马逊", "AI云"),
    "META": ("Meta", "AI平台"),
    "ORCL": ("甲骨文", "AI云"),
    "PLTR": ("Palantir", "AI软件"),
}

EXTENDED_OVERSEAS_AI = {
    "QCOM": ("高通", "边缘AI芯片"),
    "INTC": ("英特尔", "半导体"),
    "005930": ("三星电子", "半导体/存储"),
    "2454": ("联发科", "边缘AI芯片"),
    "2308": ("台达电", "数据中心电源"),
    "LITE": ("Lumentum", "光通信器件"),
    "GLW": ("康宁", "光纤/材料"),
    "SNDK": ("闪迪", "AI存储"),
    "DELL": ("戴尔科技", "AI服务器"),
}

OFFSHORE_CHINA_AI = {
    "00700": ("腾讯控股", "中国离岸AI平台/云"),
    "09988": ("阿里巴巴-W", "中国离岸AI平台/云"),
    "09888": ("百度集团-SW", "中国离岸AI/自动驾驶/云"),
    "00020": ("商汤-W", "中国离岸AI软件"),
    "09660": ("地平线机器人-W", "中国离岸AI芯片/自动驾驶"),
    "00981": ("中芯国际", "中国离岸半导体制造"),
    "01347": ("华虹半导体", "中国离岸半导体制造"),
    "03896": ("金山云", "中国离岸云计算"),
    "00992": ("联想集团", "中国离岸AI PC/服务器"),
    "01810": ("小米集团-W", "中国离岸AI终端/汽车"),
    "02228": ("晶泰控股", "中国离岸AI制药"),
}

RANKING_HEADERS = [
    "基金代码",
    "基金名称",
    "基金类型",
    "限购额度情况",
    "狭义海外AI占比",
    "扩展海外AI占比",
    "广义离岸AI占比",
    "匹配持仓数",
    "核心持仓摘要",
]

DETAIL_HEADERS = [
    "基金代码",
    "基金名称",
    "基金类型",
    "证券代码",
    "证券名称",
    "口径",
    "分类",
    "占净值比例",
    "原始重复行数",
    "原始最大行序号",
]

DICT_HEADERS = ["证券代码", "标准名称", "口径", "分类"]
PURCHASE_LIMIT_HEADERS = [
    "基金代码",
    "基金名称",
    "基金类型",
    "单位净值",
    "净值日期",
    "申购状态",
    "赎回状态",
    "下一开放日",
    "起购金额",
    "日累计限购额度",
    "日累计限购额度原始值",
    "费率",
    "申购状态代码",
    "数据源",
    "抓取时间",
]


def pct_to_float(value: str) -> float:
    value = (value or "").strip().replace(",", "")
    if not value:
        return 0.0
    if value.endswith("%"):
        value = value[:-1]
    try:
        return float(value) / 100
    except ValueError:
        return 0.0


def pct(value: float) -> str:
    return f"{value:.2%}"


def money_text(value: str) -> str:
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return "---"
    if amount < 0:
        return "---"
    if amount >= 800_000_000:
        return "无限额"
    if amount < 10_000:
        if amount.is_integer():
            return f"{int(amount)}元"
        return f"{amount:g}元"
    if amount < 100_000_000:
        amount = amount / 10_000
        return f"{amount:g}万"
    return f"{amount / 100_000_000:g}亿"


def extract_js_array(text: str, field_name: str) -> list[Any]:
    marker = f"{field_name}:"
    start = text.find(marker)
    if start < 0:
        raise ValueError(f"Cannot find {field_name} in response.")
    start = text.find("[", start)
    if start < 0:
        raise ValueError(f"Cannot find array start for {field_name}.")
    depth = 0
    in_string = False
    escape = False
    for index in range(start, len(text)):
        char = text[index]
        if in_string:
            if escape:
                escape = False
            elif char == "\\":
                escape = True
            elif char == '"':
                in_string = False
            continue
        if char == '"':
            in_string = True
        elif char == "[":
            depth += 1
        elif char == "]":
            depth -= 1
            if depth == 0:
                return ast.literal_eval(text[start : index + 1])
    raise ValueError(f"Cannot find array end for {field_name}.")


def fetch_purchase_limits() -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    params = {"t": "8", "page": "1,30000", "js": "reData", "sort": "fcode,asc"}
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://fund.eastmoney.com/Fund_sgzt.html",
    }
    last_error: Exception | None = None
    for attempt in range(4):
        try:
            response = requests.get(PURCHASE_LIMIT_URL, params=params, headers=headers, timeout=30)
            response.raise_for_status()
            text = response.text
            rows = extract_js_array(text, "datas")
            record_match = re.search(r'record:"(\d+)"', text)
            showday_match = re.search(r"showday:(\[[^\]]*\])", text)
            showdays = ast.literal_eval(showday_match.group(1)) if showday_match else []
            fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            result = {}
            snapshot_rows = []
            for item in rows:
                code = str(item[0]).zfill(6)
                limit_text = money_text(item[9])
                info = {
                    "fund_code": code,
                    "fund_name": item[1],
                    "fund_type": item[2],
                    "net_value": item[3],
                    "net_value_date": item[4],
                    "purchase_status": item[5] or "---",
                    "redemption_status": item[6] or "---",
                    "next_open_date": item[7] or "",
                    "min_purchase": money_text(item[8]),
                    "daily_limit": limit_text,
                    "daily_limit_raw": item[9],
                    "rate": item[12],
                    "purchase_status_code": item[11],
                    "source": PURCHASE_LIMIT_URL,
                    "fetched_at": fetched_at,
                }
                result[code] = info
                snapshot_rows.append(
                    [
                        code,
                        info["fund_name"],
                        info["fund_type"],
                        info["net_value"],
                        info["net_value_date"],
                        info["purchase_status"],
                        info["redemption_status"],
                        info["next_open_date"],
                        info["min_purchase"],
                        info["daily_limit"],
                        info["daily_limit_raw"],
                        info["rate"],
                        info["purchase_status_code"],
                        info["source"],
                        info["fetched_at"],
                    ]
                )
            write_csv(OUTPUT_PURCHASE_LIMIT_CSV, PURCHASE_LIMIT_HEADERS, snapshot_rows)
            metadata = {
                "purchase_limit_source": PURCHASE_LIMIT_URL,
                "purchase_limit_source_page": "https://fund.eastmoney.com/Fund_sgzt.html",
                "purchase_limit_record_count": int(record_match.group(1)) if record_match else len(rows),
                "purchase_limit_rows_fetched": len(rows),
                "purchase_limit_showdays": showdays,
                "purchase_limit_fetched_at": fetched_at,
                "purchase_limit_snapshot_csv": str(OUTPUT_PURCHASE_LIMIT_CSV),
            }
            return result, metadata
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"Failed to fetch purchase limit data: {last_error}")


def purchase_limit_summary(info: dict[str, Any] | None) -> str:
    if not info:
        return "未在天天基金申购状态总表匹配"
    status = info["purchase_status"]
    data_day = f"；数据日{info['net_value_date']}" if info.get("net_value_date") else ""
    if status == "场内交易":
        return f"场内交易；场外申购限额不适用{data_day}"
    if any(keyword in status for keyword in ["暂停", "停止", "封闭", "终止", "失败"]):
        next_open = f"；下一开放日{info['next_open_date']}" if info.get("next_open_date") else ""
        return f"{status}；当前不可场外申购{next_open}{data_day}"
    parts = [
        status,
        f"起购{info['min_purchase']}",
        f"日累计{info['daily_limit']}",
    ]
    if info.get("next_open_date"):
        parts.append(f"下一开放日{info['next_open_date']}")
    if info.get("net_value_date"):
        parts.append(f"数据日{info['net_value_date']}")
    return "；".join(parts)


def classify(code: str) -> tuple[str, str] | None:
    if code in STRICT_OVERSEAS_AI:
        return "狭义海外AI", STRICT_OVERSEAS_AI[code][1]
    if code in EXTENDED_OVERSEAS_AI:
        return "扩展海外AI", EXTENDED_OVERSEAS_AI[code][1]
    if code in OFFSHORE_CHINA_AI:
        return "广义离岸中国AI", OFFSHORE_CHINA_AI[code][1]
    return None


def read_deduped_holdings() -> tuple[dict[tuple[str, str], dict[str, Any]], int, int]:
    raw_rows = 0
    holdings: dict[tuple[str, str], dict[str, Any]] = {}
    with INPUT.open("r", newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            raw_rows += 1
            fund_code = row["基金代码"]
            security_code = row["证券代码"].upper().strip()
            value = pct_to_float(row["占净值比例"])
            key = (fund_code, security_code)
            previous = holdings.get(key)
            if previous is None or value > previous["ratio"]:
                holdings[key] = {
                    "fund_code": fund_code,
                    "fund_name": row["基金名称"],
                    "fund_type": row["基金类型"],
                    "security_code": security_code,
                    "security_name": row["证券名称"],
                    "ratio": value,
                    "rank": row["序号"],
                    "duplicate_rows": 1 if previous is None else previous["duplicate_rows"] + 1,
                }
            else:
                previous["duplicate_rows"] += 1
    return holdings, raw_rows, raw_rows - len(holdings)


def analyze() -> tuple[list[list[Any]], list[list[Any]], list[list[Any]], dict[str, Any]]:
    purchase_limits, purchase_metadata = fetch_purchase_limits()
    holdings, raw_rows, duplicate_rows_removed = read_deduped_holdings()
    fund_records: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "fund_name": "",
            "fund_type": "",
            "strict": 0.0,
            "extended": 0.0,
            "offshore_china": 0.0,
            "details": [],
        }
    )
    detail_rows: list[list[Any]] = []
    dictionary_rows: list[list[Any]] = []

    for code, (name, category) in STRICT_OVERSEAS_AI.items():
        dictionary_rows.append([code, name, "狭义海外AI", category])
    for code, (name, category) in EXTENDED_OVERSEAS_AI.items():
        dictionary_rows.append([code, name, "扩展海外AI", category])
    for code, (name, category) in OFFSHORE_CHINA_AI.items():
        dictionary_rows.append([code, name, "广义离岸中国AI", category])

    for item in holdings.values():
        classified = classify(item["security_code"])
        if not classified:
            continue
        bucket, category = classified
        fund = fund_records[item["fund_code"]]
        fund["fund_name"] = item["fund_name"]
        fund["fund_type"] = item["fund_type"]
        if bucket == "狭义海外AI":
            fund["strict"] += item["ratio"]
        elif bucket == "扩展海外AI":
            fund["extended"] += item["ratio"]
        else:
            fund["offshore_china"] += item["ratio"]
        fund["details"].append((item["ratio"], item["security_code"], item["security_name"], bucket, category))
        detail_rows.append(
            [
                item["fund_code"],
                item["fund_name"],
                item["fund_type"],
                item["security_code"],
                item["security_name"],
                bucket,
                category,
                item["ratio"],
                item["duplicate_rows"],
                item["rank"],
            ]
        )

    ranking_rows: list[list[Any]] = []
    for fund_code, fund in fund_records.items():
        strict = fund["strict"]
        extended_total = fund["strict"] + fund["extended"]
        broad = extended_total + fund["offshore_china"]
        top = "; ".join(
            f"{name} {pct(value)}"
            for value, _code, name, _bucket, _category in sorted(fund["details"], reverse=True)[:8]
        )
        ranking_rows.append(
            [
                fund_code,
                fund["fund_name"],
                fund["fund_type"],
                purchase_limit_summary(purchase_limits.get(fund_code)),
                strict,
                extended_total,
                broad,
                len(fund["details"]),
                top,
            ]
        )

    ranking_rows.sort(key=lambda row: row[4], reverse=True)
    detail_rows.sort(key=lambda row: (row[0], -row[7], row[3]))
    dictionary_rows.sort(key=lambda row: (row[2], row[0]))
    summary = {
        "input": str(INPUT),
        "raw_stock_rows": raw_rows,
        "deduped_fund_security_rows": len(holdings),
        "duplicate_rows_removed_by_fund_security": duplicate_rows_removed,
        "funds_with_any_ai_match": len(ranking_rows),
        "funds_with_strict_overseas_ai": sum(1 for row in ranking_rows if row[4] > 0),
        "funds_with_extended_overseas_ai": sum(1 for row in ranking_rows if row[5] > 0),
        "funds_with_broad_offshore_ai": sum(1 for row in ranking_rows if row[6] > 0),
        "purchase_limit_matched_funds": sum(
            1 for row in ranking_rows if row[3] != "未在天天基金申购状态总表匹配"
        ),
        "purchase_limit_unmatched_funds": sum(
            1 for row in ranking_rows if row[3] == "未在天天基金申购状态总表匹配"
        ),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    summary.update(purchase_metadata)
    return ranking_rows, detail_rows, dictionary_rows, summary


def write_csv(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([pct(value) if isinstance(value, float) else value for value in row])


def style_header(ws: Any) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(name="Arial", bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
    ws.freeze_panes = "A2"


def append_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]], limit: int | None = None) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows[:limit] if limit else rows:
        ws.append(row)
    style_header(ws)
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200])
        width = min(max(max_length + 2, 10), 55)
        ws.column_dimensions[column_cells[0].column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00%"


def write_workbook(ranking_rows: list[list[Any]], detail_rows: list[list[Any]], dictionary_rows: list[list[Any]], summary: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "方法说明"
    ws.append(["项目", "内容"])
    notes = [
        ["分析口径", f"基于 {QUARTER.source_stock_csv_relative.as_posix()}；同一基金同一证券代码重复出现时取最大占净值比例。"],
        ["狭义海外AI占比", "只统计美股/台股/韩股/欧洲等海外AI核心算力、AI云、大模型平台、核心半导体和数据中心基础设施。"],
        ["扩展海外AI占比", "狭义海外AI + 海外AI相关但纯度较低的服务器、光通信、边缘芯片、存储等标的。"],
        ["广义离岸AI占比", "扩展海外AI + 港股/离岸中国AI平台、云、半导体、AI软件、AI终端等标的。"],
        ["未纳入口径", "苹果、特斯拉、美团、快手、网易等未计入狭义海外AI；如需泛科技口径可另做一版。"],
        ["生成摘要", json.dumps(summary, ensure_ascii=False, sort_keys=True)],
    ]
    for row in notes:
        ws.append(row)
    style_header(ws)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 120

    strict_rows = [row for row in ranking_rows if row[4] > 0]
    extended_rows = sorted([row for row in ranking_rows if row[5] > 0], key=lambda row: row[5], reverse=True)
    broad_rows = sorted([row for row in ranking_rows if row[6] > 0], key=lambda row: row[6], reverse=True)

    append_sheet(wb, "狭义海外AI排名", RANKING_HEADERS, strict_rows)
    append_sheet(wb, "扩展海外AI排名", RANKING_HEADERS, extended_rows)
    append_sheet(wb, "广义离岸AI排名", RANKING_HEADERS, broad_rows)
    append_sheet(wb, "匹配持仓明细", DETAIL_HEADERS, detail_rows)
    append_sheet(wb, "标的字典", DICT_HEADERS, dictionary_rows)
    wb.save(OUTPUT_XLSX)


def validate_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return {"opened": True, "sheetnames": wb.sheetnames, "sheet_count": len(wb.sheetnames)}
    finally:
        wb.close()


def main() -> int:
    ranking_rows, detail_rows, dictionary_rows, summary = analyze()
    write_csv(OUTPUT_RANKING_CSV, RANKING_HEADERS, ranking_rows)
    write_csv(OUTPUT_DETAIL_CSV, DETAIL_HEADERS, detail_rows)
    write_workbook(ranking_rows, detail_rows, dictionary_rows, summary)
    summary["outputs"] = {
        "workbook": str(OUTPUT_XLSX),
        "ranking_csv": str(OUTPUT_RANKING_CSV),
        "detail_csv": str(OUTPUT_DETAIL_CSV),
    }
    summary["workbook_validation"] = validate_workbook(OUTPUT_XLSX)
    OUTPUT_SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
