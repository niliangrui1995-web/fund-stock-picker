from __future__ import annotations

import csv
import io
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from spreadsheet_safety import append_safe_xlsx_row, safe_csv_row


class SpreadsheetSafetyTests(unittest.TestCase):
    def test_csv_formula_prefixes_are_neutralized_once(self) -> None:
        unsafe_values = ["=1+1", "+SUM(A1:A2)", "-1+1", "@SUM(A1:A2)", "＝1+1", "\t=1+1", "\x00=1+1"]
        escaped = safe_csv_row(unsafe_values)

        self.assertTrue(all(isinstance(value, str) and value.startswith("\t") for value in escaped))
        self.assertEqual(escaped[0], "\t=1+1")
        self.assertEqual(escaped[5], "\t=1+1")
        self.assertEqual(escaped[6], "\t=1+1")
        self.assertEqual(safe_csv_row(escaped), escaped)
        self.assertEqual(safe_csv_row(["普通文本", "  正常", 42]), ["普通文本", "  正常", 42])

        output = io.StringIO(newline="")
        csv.writer(output, quoting=csv.QUOTE_ALL).writerow(escaped)
        self.assertIn('"\t=1+1"', output.getvalue())

    def test_xlsx_formula_prefixes_remain_string_cells(self) -> None:
        values = ["=1+1", "\t=1+1", "\u200b+SUM(A1:A2)", "--", "普通文本", 42]
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "safety.xlsx"

            workbook = Workbook()
            append_safe_xlsx_row(workbook.active, values)
            workbook.save(path)
            workbook.close()

            loaded = load_workbook(path, data_only=False, read_only=False)
            worksheet = loaded.active
            self.assertEqual(worksheet["A1"].value, "=1+1")
            self.assertEqual(worksheet["B1"].value, "=1+1")
            self.assertEqual(worksheet["C1"].value, "+SUM(A1:A2)")
            self.assertEqual(worksheet["D1"].value, "--")
            self.assertTrue(all(worksheet.cell(1, column).data_type != "f" for column in range(1, 7)))
            loaded.close()

    def test_write_only_workbook_forces_string_cells(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "write-only-safety.xlsx"

            workbook = Workbook(write_only=True)
            worksheet = workbook.create_sheet("数据")
            append_safe_xlsx_row(worksheet, ["=1+1", "\t=1+1"])
            workbook.save(path)

            loaded = load_workbook(path, data_only=False, read_only=False)
            self.assertTrue(all(loaded["数据"].cell(1, column).data_type != "f" for column in range(1, 3)))
            loaded.close()


if __name__ == "__main__":
    unittest.main()
